import pandas as pd
import numpy as np
from sklearn.preprocessing import LabelEncoder,OneHotEncoder
import warnings
warnings.filterwarnings('ignore')
from faker import Faker
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import pickle
import joblib
from sklearn.decomposition import PCA
from sklearn.model_selection import GridSearchCV
import openpyxl
import matplotlib.pyplot as plt
from sklearn.metrics import ConfusionMatrixDisplay, roc_curve, auc
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix
from sklearn.ensemble import RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
fake = Faker()

# -------------------------------
# Risk score calculation
# -------------------------------
def calculate_risk(row):
    risk = 0
    risk += (row['Age'] / 85) * 20  # age
    if row['Cholesterol'] > 240: 
        risk += 20
    elif row['Cholesterol'] > 200: 
        risk += 10
    if row['Blood_Sugar'] > 126: 
        risk += 15
    elif row['Blood_Sugar'] > 100: 
        risk += 8
    if row['Smoking'] == 'Yes': 
        risk += 15
    if row['Family_History_Cancer'] == 'Yes': 
        risk += 20
    risk += min(row['Tumor_Size_cm'] * 3, 15)  # tumor size
    return min(risk, 100)

# -------------------------------
# Add anomalies
# -------------------------------
def add_anomalies(data, fraction_nulls=0.02, fraction_duplicates=0.02,
                  fraction_outliers=0.02, fraction_incorrect=0.02,
                  random_state=42):
    np.random.seed(random_state)
    n_rows = len(data)
    numeric_cols = ['Cholesterol', 'Blood_Sugar', 'BMI', 'Tumor_Size_cm']

    # Nulls
    null_indices = np.random.choice(data.index, int(n_rows * fraction_nulls), replace=False)
    for idx in null_indices:
        col = np.random.choice(data.columns)
        data.loc[idx, col] = np.nan

    # Duplicates
    duplicate_indices = np.random.choice(data.index, int(n_rows * fraction_duplicates), replace=False)
    duplicates = data.loc[duplicate_indices].copy()
    data = pd.concat([data, duplicates], ignore_index=True)

    # Outliers (multiply some values by big factor)
    outlier_indices = np.random.choice(data.index, int(n_rows * fraction_outliers), replace=False)
    for col in numeric_cols:
        data.loc[outlier_indices, col] *= np.random.uniform(3, 10, size=len(outlier_indices))

    # Incorrect formats(Some numeric fields (Cholesterol, Blood_Sugar) are replaced with "abc" → simulates dirty text in numbers.)
    incorrect_indices = np.random.choice(data.index, int(n_rows * fraction_incorrect), replace=False)
    for col in ['Cholesterol', 'Blood_Sugar']:
        data[col] = data[col].astype(object)
        data.loc[incorrect_indices, col] = "abc"

    return data

# -------------------------------
# Generate dataset
# -------------------------------
def generate_data(n_samples=1000, random_state=42):
    np.random.seed(random_state)
    data = pd.DataFrame({
        'Patient_ID': range(1, n_samples + 1),
        'Name': [fake.name() for _ in range(n_samples)],
        'Age': np.random.randint(18, 85, size=n_samples).round(0),
        'Gender': np.random.choice(['Male', 'Female'], size=n_samples, p=[0.48, 0.52]),
        'Height_cm': np.random.normal(170, 10, size=n_samples).round(0),
        'Weight_kg': np.random.normal(75, 15, size=n_samples).round(0),
        'BMI': np.random.uniform(18, 35, size=n_samples).round(1),
        'Systolic_BP': np.random.normal(120, 20, size=n_samples).round(1),
        'Diastolic_BP': np.random.normal(80, 15, size=n_samples).round(1),
        'Heart_Rate': np.random.normal(70, 12, size=n_samples).round(1),
        'Temperature_F': np.random.normal(98.6, 1.5, size=n_samples).round(1),
        'Blood_Sugar': np.random.normal(100, 30, size=n_samples).round(1),
        'Cholesterol': np.random.normal(200, 40, size=n_samples).round(1),
        'Hemoglobin': np.random.normal(14, 2, size=n_samples).round(1),
        'Smoking': np.random.choice(['Yes', 'No'], size=n_samples, p=[0.25, 0.75]),
        'Alcohol_Consumption': np.random.choice(['Yes', 'No'], size=n_samples, p=[0.3, 0.7]),
        'Exercise_Hours_Week': np.random.poisson(3, size=n_samples),
        'Diabetes': np.random.choice(['Yes', 'No'], size=n_samples, p=[0.15, 0.85]),
        'Hypertension': np.random.choice(['Yes', 'No'], size=n_samples, p=[0.20, 0.80]),
        'Heart_Disease': np.random.choice(['Yes', 'No'], size=n_samples, p=[0.10, 0.90]),
        'Hospital_Visits_Year': np.random.poisson(2, size=n_samples),
        'Insurance_Type': np.random.choice(['Public', 'Private', 'Uninsured'], size=n_samples, p=[0.4, 0.5, 0.1]),
        'Family_History_Cancer': np.random.choice(['Yes', 'No'], size=n_samples, p=[0.18, 0.82]),
        'Biopsy_Performed': np.random.choice(['Yes', 'No'], size=n_samples, p=[0.06, 0.94]),
        'Biopsy_Result': np.random.choice(['Benign', 'Malignant', 'Unknown'], size=n_samples, p=[0.85, 0.10, 0.05]),
        'Treatment': np.random.choice(['Chemotherapy', 'Radiation Therapy', 'Surgery', 'Immunotherapy', 'None'],
                                      size=n_samples, p=[0.2, 0.15, 0.25, 0.1, 0.3]),
        'Allergies': np.random.choice(['None', 'Penicillin', 'Peanuts', 'Seafood', 'Dust', 'Pollen'],
                                      size=n_samples, p=[0.5, 0.15, 0.1, 0.1, 0.08, 0.07]),
        'Sonography_Result': np.random.choice(['Normal', 'Abnormal'], size=n_samples, p=[0.85, 0.15]),
        'Tumor_Size_cm': np.round(np.random.uniform(0, 10, size=n_samples), 2)
    })

    data['BMI'] = (data['Weight_kg'] / ((data['Height_cm'] / 100) ** 2)).round(1)
    data['Risk_Score'] = data.apply(calculate_risk, axis=1).round(1)
    return add_anomalies(data, random_state=random_state)

# -------------------------------
# Cleaning with IQR + Z-Score + Rules
# -------------------------------
def clean_dataset(df, z_thresh=3):
    print("Initial shape:", df.shape)

    # 1. Handle Missing(nulls)
    df = df.dropna()

    # 2. Remove Duplicates
    df = df.drop_duplicates()

    # 3. Fix incorrect data types
    for col in ['Cholesterol', 'Blood_Sugar']:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=['Cholesterol', 'Blood_Sugar'])

    # 4. Outlier detection using IQR
    numeric_cols = ['Cholesterol', 'Blood_Sugar', 'BMI', 'Tumor_Size_cm']
    for col in numeric_cols:
        Q1, Q3 = df[col].quantile([0.25, 0.75])
        IQR = Q3 - Q1
        lower, upper = Q1 - 1.5 * IQR, Q3 + 1.5 * IQR
        df = df[(df[col] >= lower) & (df[col] <= upper)]

    # 5. Outlier detection using Z-score
    for col in numeric_cols:
        z_scores = (df[col] - df[col].mean()) / df[col].std()
        df = df[np.abs(z_scores) < z_thresh]

    # 6. Domain rule filtering
    df = df[(df['Age'] >= 0) & (df['Age'] <= 120)]
    df = df[(df['BMI'] >= 10) & (df['BMI'] <= 60)]
    df = df[(df['Blood_Sugar'] >= 40) & (df['Blood_Sugar'] <= 500)]
    df = df[(df['Cholesterol'] >= 80) & (df['Cholesterol'] <= 400)]

    print("Cleaned shape:", df.shape)
    return df

# -------------------------------
# Run full pipeline
# -------------------------------
# -------------------------------
if __name__ == "__main__":
    TARGET_ROWS = 1000

    print("Generating dataset with anomalies...")
    dataset = generate_data(TARGET_ROWS)
    numeric_cols = dataset.select_dtypes(include=[np.number]).columns
    dataset[numeric_cols] = dataset[numeric_cols].round(2)
    dataset.to_csv("patient_data_with_anomalies.csv", index=False)
    dataset.to_excel("patient_data_with_anomalies.xlsx", index=False)  # <-- This is the Excel export

    print("Cleaning dataset (IQR + Z-Score + Rules)...")
    cleaned = clean_dataset(dataset)

    # If rows are missing after cleaning
    if len(cleaned) < TARGET_ROWS:
        extra_needed = TARGET_ROWS - len(cleaned)
        print(f"⚠️ Only {len(cleaned)} rows after cleaning. Generating {extra_needed*2} more...")

        # Generate a bigger batch of extra data
        extra_data = generate_data(extra_needed * 2, random_state=np.random.randint(0, 99999))
        extra_cleaned = clean_dataset(extra_data)

        # Merge original + extra
        cleaned = pd.concat([cleaned, extra_cleaned], ignore_index=True)

    #  Ensure exactly 1000 rows (cut off extras if too many)
    cleaned = cleaned.head(TARGET_ROWS)

    #  ROUND numeric columns to 2 decimal places
    numeric_cols = cleaned.select_dtypes(include=[np.number]).columns
    cleaned[numeric_cols] = cleaned[numeric_cols].round(2)

    print(f"✅ Final dataset has exactly {len(cleaned)} rows")

    #  SAVE to CSV and Excel
    cleaned.to_csv("patient_data_cleaned.csv", index=False)
    cleaned.to_excel("patient_data_cleaned.xlsx", index=False)  # <-- This is the Excel export

    print("Saved cleaned dataset to CSV and Excel.")

# -------------------------------
# Data Integrity Checker
# -------------------------------
def check_data_integrity(df):
    issues = []

    # Convert to numeric safely
    for col in ['Cholesterol', 'Blood_Sugar']:
        df[col] = pd.to_numeric(df[col], errors='coerce')  # turns "abc" into NaN

    # 1. Check data types
    expected_types = {
        'Age': (int, float),
        'Cholesterol': (int, float),
        'Blood_Sugar': (int, float),
        'BMI': (int, float),
        'Tumor_Size_cm': (int, float),
        'Smoking': str,
        'Family_History_Cancer': str
    }

    for col, typ in expected_types.items():
        if not df[col].map(lambda x: isinstance(x, typ)).all():
            issues.append(f"Invalid data types in column: {col}")

    # 2. Check value ranges
    if df['Age'].dropna().lt(0).any() or df['Age'].dropna().gt(120).any():
        issues.append("Age out of expected range (0-120)")
    if df['BMI'].dropna().lt(10).any() or df['BMI'].dropna().gt(60).any():
        issues.append("BMI out of expected range (10-60)")
    if df['Blood_Sugar'].dropna().lt(40).any() or df['Blood_Sugar'].dropna().gt(500).any():
        issues.append("Blood Sugar out of expected range (40-500)")
    if df['Cholesterol'].dropna().lt(80).any() or df['Cholesterol'].dropna().gt(400).any():
        issues.append("Cholesterol out of expected range (80-400)")
    if df['Tumor_Size_cm'].dropna().lt(0).any() or df['Tumor_Size_cm'].dropna().gt(20).any():
        issues.append("Tumor size out of range (0-20)")

    # 3. Check categorical values
    for col, allowed in {
        'Smoking': ['Yes', 'No'],
        'Family_History_Cancer': ['Yes', 'No'],
        'Gender': ['Male', 'Female']
    }.items():
        if not df[col].isin(allowed).all():
            issues.append(f"Unexpected values in column: {col}")

    # 4. Logical consistency
    if 'Systolic_BP' in df.columns and 'Diastolic_BP' in df.columns:
        inconsistent_bp = (df['Systolic_BP'] < df['Diastolic_BP']).sum()
        if inconsistent_bp > 0:
            issues.append(f"{inconsistent_bp} rows where Diastolic BP > Systolic BP")

    # Output results
    if issues:
        print("❌ Data Integrity Issues Found:")
        with open("data_integrity_issues.txt", "w") as f:
            for i in issues:
                print(f" - {i}")
                f.write(f"{i}\n")
        print("📄 Saved issues to data_integrity_issues.txt")

    else:
        print("✅ Data Integrity Check Passed")


# -------------------------------
# Run full pipeline with integrity checks
# -------------------------------
if __name__ == "__main__":
    print("Generating dataset with anomalies...")
    dataset = generate_data(1000)

    # print("\n🔎 Checking integrity BEFORE cleaning...")
    # check_data_integrity(dataset)

    # Save dataset with anomalies
    # Save dataset with anomalies
    dataset.to_csv("patient_data_with_anomalies.csv", index=False)
    print(" Saved: patient_data_with_anomalies.csv")

# Save cleaned dataset
    cleaned.to_csv("patient_data_cleaned.csv", index=False)
    print(" Saved: patient_data_cleaned.csv")

# Save integrity-checked dataset



    # Final integrity check - save only if passed
    print("\n🧪 Saving only integrity-passed data...")
    def passes_integrity(df):
        issues = []
        # Convert columns safely
        for col in ['Cholesterol', 'Blood_Sugar']:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        # Check for NaNs
        if df[['Cholesterol', 'Blood_Sugar']].isna().any().any():
            return False
        if df['Age'].between(0, 120).all() and \
           df['BMI'].between(10, 60).all() and \
           df['Blood_Sugar'].between(40, 500).all() and \
           df['Cholesterol'].between(80, 400).all():
            return True
        return False

    if passes_integrity(cleaned):
        cleaned.to_csv("patient_data_integrity_checked.csv", index=False)
        print(" Saved: patient_data_integrity_checked.csv")
    else:
        print("Integrity check failed. File not saved.")



# pd.get_dummies()

def prepare_features_and_target(data):
    processed_data = data.copy()

    # Create target variable based on risk assessment
    risk_threshold = processed_data['Risk_Score'].median() #calculate median of risk_score
    processed_data['High_Risk'] = (processed_data['Risk_Score'] > risk_threshold).astype(int)   #if risk_score is > threshold then it considerd as high risk factor

    # Encode categorical variables
    label_encoders = {}
    categorical_columns = ['Gender', 'Smoking', 'Alcohol_Consumption', 'Diabetes',
                          'Hypertension', 'Heart_Disease', 'Insurance_Type']

    for col in categorical_columns:
        if col in processed_data.columns:
            le = LabelEncoder()
            processed_data[col + '_encoded'] = le.fit_transform(processed_data[col])
            label_encoders[col] = le

    # Select feature columns
    feature_columns = [
        'Age', 'Height_cm', 'Weight_kg', 'BMI', 'Systolic_BP', 'Diastolic_BP',
        'Heart_Rate', 'Temperature_F', 'Blood_Sugar', 'Cholesterol', 'Hemoglobin',
        'Exercise_Hours_Week', 'Hospital_Visits_Year', 'Gender_encoded',
        'Smoking_encoded', 'Alcohol_Consumption_encoded', 'Diabetes_encoded',
        'Hypertension_encoded', 'Heart_Disease_encoded', 'Insurance_Type_encoded'
    ]

    # Filter existing columns
    available_features = [col for col in feature_columns if col in processed_data.columns]

    X = processed_data[available_features]
    y = processed_data['High_Risk']

    return X, y, label_encoders, processed_data

def save_feature_info(X, y, label_encoders):
    feature_info = {
        'feature_names': list(X.columns),
        'target_name': 'High_Risk',
        'n_features': X.shape[1],
        'n_samples': X.shape[0],
        'target_distribution': y.value_counts().to_dict()
    }

    # Save encoders and feature info
    import pickle
    with open('label_encoders.pkl', 'wb') as f:
        pickle.dump(label_encoders, f)

    with open('feature_info.pkl', 'wb') as f:
        pickle.dump(feature_info, f)

    return feature_info

if __name__ == "__main__":
    print("Preparing features and target variable...")
    data = pd.read_csv('patient_data_integrity_checked.csv')

    X, y, label_encoders, processed_data = prepare_features_and_target(data)
    feature_info = save_feature_info(X, y, label_encoders)

    # Save features and target
    X.to_csv('features.csv', index=False)
    y.to_csv('target.csv', index=False)
    processed_data.to_csv('processed_data.csv', index=False)

    print(f"Features shape: {X.shape}")
    print(f"Target shape: {y.shape}")
    print(f"Target distribution: {feature_info['target_distribution']}")
    print(f"Feature columns: {len(feature_info['feature_names'])}")
    print("Saved files: features.csv, target.csv, processed_data.csv")
    print("Saved encoders: label_encoders.pkl, feature_info.pkl")




def scale_and_split_data(X, y, test_size=0.2, random_state=42):
    # Train-test split
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=test_size, random_state=random_state, stratify=y
    )

    # Feature scaling
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)

    # Dimensionality Reduction with PCA
    pca = PCA(n_components=5, random_state=random_state)
    X_train_pca = pca.fit_transform(X_train_scaled)
    X_test_pca = pca.transform(X_test_scaled)

    # Convert back to DataFrames
    X_train_pca_df = pd.DataFrame(X_train_pca, columns=[f'PC{i+1}' for i in range(X_train_pca.shape[1])])
    X_test_pca_df = pd.DataFrame(X_test_pca, columns=[f'PC{i+1}' for i in range(X_test_pca.shape[1])])

    return X_train_pca_df, X_test_pca_df, y_train, y_test, scaler, pca

def save_split_data(X_train, X_test, y_train, y_test, scaler):
    # Save datasets
    X_train.to_csv('X_train.csv', index=False)
    X_test.to_csv('X_test.csv', index=False)
    y_train.to_csv('y_train.csv', index=False)
    y_test.to_csv('y_test.csv', index=False)

    # Save scaler
    with open('scaler.pkl', 'wb') as f:
        pickle.dump(scaler, f)

    split_info = {
        'train_size': X_train.shape[0],
        'test_size': X_test.shape[0],
        'train_ratio': X_train.shape[0] / (X_train.shape[0] + X_test.shape[0]),
        'test_ratio': X_test.shape[0] / (X_train.shape[0] + X_test.shape[0]),
        'train_target_dist': y_train.value_counts().to_dict(),
        'test_target_dist': y_test.value_counts().to_dict()
    }

    with open('split_info.pkl', 'wb') as f:
        pickle.dump(split_info, f)

    return split_info

if __name__ == "__main__":
    print("Loading features and target, then scaling and splitting...")
    X = pd.read_csv('features.csv')
    y = pd.read_csv('target.csv').squeeze()

    X_train, X_test, y_train, y_test, scaler, pca = scale_and_split_data(X, y)
    split_info = save_split_data(X_train, X_test, y_train, y_test, scaler)

    print(f"Training set: {split_info['train_size']} samples ({split_info['train_ratio']:.2%})")
    print(f"Test set: {split_info['test_size']} samples ({split_info['test_ratio']:.2%})")
    print(f"Train target distribution: {split_info['train_target_dist']}")
    print(f"Test target distribution: {split_info['test_target_dist']}")
    print("Saved files: X_train.csv, X_test.csv, y_train.csv, y_test.csv")
    print("Saved objects: scaler.pkl, split_info.pkl")

    # Hyperparameter tuning for Logistic Regression
    print("\n🔁 Tuning Logistic Regression model...")
    param_grid_lr = {'C': [0.01, 0.1, 1, 10], 'solver': ['lbfgs', 'liblinear']}
    grid_lr = GridSearchCV(LogisticRegression(max_iter=1000, random_state=42), param_grid_lr, cv=3, scoring='accuracy')
    grid_lr.fit(X_train, y_train)
    model = grid_lr.best_estimator_
    print(f"Best Logistic Regression Params: {grid_lr.best_params_}")

    # -------------------------
    # 🧪 Evaluate the Model
    # -------------------------
    print("\n📊 Evaluation on Test Set:")
    y_pred = model.predict(X_test)
    y_proba = model.predict_proba(X_test)[:, 1]

    acc = accuracy_score(y_test, y_pred)
    cm = confusion_matrix(y_test, y_pred)
    report = classification_report(y_test, y_pred)

    print(f"Accuracy: {acc:.4f}")
    print("Confusion Matrix:")
    print(cm)
    print("\nClassification Report:")
    print(report)
    # Save predictions to CSV
    pd.DataFrame({'y_true': y_test, 'y_pred_logistic': y_pred, 'y_proba_logistic': y_proba}).to_csv('logistic_predictions.csv', index=False)

        # Visualization: Confusion Matrix and ROC Curve

        # Confusion Matrix
    disp = ConfusionMatrixDisplay(confusion_matrix=cm)
    disp.plot(cmap=plt.cm.Blues)
    plt.title('Logistic Regression Confusion Matrix')
    plt.savefig('logistic_confusion_matrix.png')
    plt.close()

    # ROC Curve
    fpr, tpr, _ = roc_curve(y_test, y_proba)
    roc_auc = auc(fpr, tpr)
    plt.figure()
    plt.plot(fpr, tpr, color='darkorange', lw=2, label=f'ROC curve (area = {roc_auc:.2f})')
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Logistic Regression ROC Curve')
    plt.legend(loc="lower right")
    plt.savefig('logistic_roc_curve.png')
    plt.close()

    # Line plot of predicted probabilities
    plt.figure(figsize=(10,4))
    plt.plot(range(len(y_proba)), y_proba, marker='o', linestyle='-', color='blue')
    plt.title('Predicted Probability of Cancer Cell Presence (Logistic Regression)')
    plt.xlabel('Sample Index')
    plt.ylabel('Predicted Probability')
    plt.tight_layout()
    plt.savefig('logistic_predicted_probabilities.png')
    plt.close()

    # Export line plot to Excel
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    wb = Workbook()
    ws = wb.active
    ws.title = "Logistic Regression Prediction"
    ws['A1'] = 'Predicted Probability of Cancer Cell Presence'
    img = XLImage('logistic_predicted_probabilities.png')
    ws.add_image(img, 'B3')
    wb.save('logistic_prediction_plot.xlsx')

    # -------------------------
    # 💾 Save the Model
    # -------------------------
    joblib.dump(model, 'logistic_model.pkl')
    print("✅ Model saved as logistic_model.pkl")

    # -------------------------
    # 🌲 Train Random Forest Model
    # -------------------------
    print("\n🌲 Training Random Forest model...")
    # Hyperparameter tuning for Random Forest
    print("\n🌲 Tuning Random Forest model...")
    param_grid_rf = {'n_estimators': [50, 100], 'max_depth': [5, 10, None]}
    grid_rf = GridSearchCV(RandomForestClassifier(random_state=42), param_grid_rf, cv=3, scoring='accuracy')
    grid_rf.fit(X_train, y_train)
    rf_model = grid_rf.best_estimator_
    print(f"Best Random Forest Params: {grid_rf.best_params_}")

    # -------------------------
    # 🧪 Evaluate Random Forest
    # -------------------------
    print("\n📊 Evaluation of Random Forest on Test Set:")
    rf_pred = rf_model.predict(X_test)
    rf_proba = rf_model.predict_proba(X_test)[:, 1]

    rf_acc = accuracy_score(y_test, rf_pred)
    rf_cm = confusion_matrix(y_test, rf_pred)
    rf_report = classification_report(y_test, rf_pred)

    print(f"Random Forest Accuracy: {rf_acc:.4f}")
    print("Confusion Matrix:")
    print(rf_cm)
    print("\nClassification Report:")
    print(rf_report)
    # Save predictions to CSV
    pd.DataFrame({'y_true': y_test, 'y_pred_rf': rf_pred, 'y_proba_rf': rf_proba}).to_csv('random_forest_predictions.csv', index=False)

        # Visualization: Confusion Matrix and ROC Curve
    disp_rf = ConfusionMatrixDisplay(confusion_matrix=rf_cm)
    disp_rf.plot(cmap=plt.cm.Blues)
    plt.title('Random Forest Confusion Matrix')
    plt.savefig('random_forest_confusion_matrix.png')
    plt.close()

    fpr_rf, tpr_rf, _ = roc_curve(y_test, rf_proba)
    roc_auc_rf = auc(fpr_rf, tpr_rf)
    plt.figure()
    plt.plot(fpr_rf, tpr_rf, color='darkorange', lw=2, label=f'ROC curve (area = {roc_auc_rf:.2f})')
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Random Forest ROC Curve')
    plt.legend(loc="lower right")
    plt.savefig('random_forest_roc_curve.png')
    plt.close()

    # Line plot of predicted probabilities
    plt.figure(figsize=(10,4))
    plt.plot(range(len(rf_proba)), rf_proba, marker='o', linestyle='-', color='green')
    plt.title('Predicted Probability of Cancer Cell Presence (Random Forest)')
    plt.xlabel('Sample Index')
    plt.ylabel('Predicted Probability')
    plt.tight_layout()
    plt.savefig('random_forest_predicted_probabilities.png')
    plt.close()

    # Export line plot to Excel
    wb_rf = Workbook()
    ws_rf = wb_rf.active
    ws_rf.title = "Random Forest Prediction"
    ws_rf['A1'] = 'Predicted Probability of Cancer Cell Presence'
    img_rf = XLImage('random_forest_predicted_probabilities.png')
    ws_rf.add_image(img_rf, 'B3')
    wb_rf.save('random_forest_prediction_plot.xlsx')

    # -------------------------
    # 💾 Save Random Forest Model
    # -------------------------
    joblib.dump(rf_model, 'random_forest_model.pkl')
    print("✅ Random Forest model saved as random_forest_model.pkl")
    

    # -------------------------------
    # Train and Evaluate Decision Tree
    # -------------------------------
    print("\n🌳 Training Decision Tree model...")

    # Initialize and train
    # Hyperparameter tuning for Decision Tree
    print("\n🌳 Tuning Decision Tree model...")
    param_grid_dt = {'max_depth': [3, 5, 10, None], 'min_samples_split': [2, 5, 10]}
    grid_dt = GridSearchCV(DecisionTreeClassifier(random_state=42), param_grid_dt, cv=3, scoring='accuracy')
    grid_dt.fit(X_train, y_train)
    dt_model = grid_dt.best_estimator_
    print(f"Best Decision Tree Params: {grid_dt.best_params_}")

    # Predict on test set
    y_pred_dt = dt_model.predict(X_test)
    dt_proba = dt_model.predict_proba(X_test)[:, 1]

    # Evaluate
    dt_accuracy = accuracy_score(y_test, y_pred_dt)
    dt_conf_matrix = confusion_matrix(y_test, y_pred_dt)
    dt_report = classification_report(y_test, y_pred_dt)

    # Display results
    print("\n📊 Evaluation of Decision Tree on Test Set:")
    print(f"Decision Tree Accuracy: {dt_accuracy:.4f}")
    print("Confusion Matrix:")
    print(dt_conf_matrix)
    print("\nClassification Report:")
    print(dt_report)
    # Save predictions to CSV
    pd.DataFrame({'y_true': y_test, 'y_pred_dt': y_pred_dt, 'y_proba_dt': dt_proba}).to_csv('decision_tree_predictions.csv', index=False)

        # Visualization: Confusion Matrix and ROC Curve
    disp_dt = ConfusionMatrixDisplay(confusion_matrix=dt_conf_matrix)
    disp_dt.plot(cmap=plt.cm.Blues)
    plt.title('Decision Tree Confusion Matrix')
    plt.savefig('decision_tree_confusion_matrix.png')
    plt.close()

    fpr_dt, tpr_dt, _ = roc_curve(y_test, dt_proba)
    roc_auc_dt = auc(fpr_dt, tpr_dt)
    plt.figure()
    plt.plot(fpr_dt, tpr_dt, color='darkorange', lw=2, label=f'ROC curve (area = {roc_auc_dt:.2f})')
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Decision Tree ROC Curve')
    plt.legend(loc="lower right")
    plt.savefig('decision_tree_roc_curve.png')
    plt.close()

    # Line plot of predicted probabilities
    plt.figure(figsize=(10,4))
    plt.plot(range(len(dt_proba)), dt_proba, marker='o', linestyle='-', color='red')
    plt.title('Predicted Probability of Cancer Cell Presence (Decision Tree)')
    plt.xlabel('Sample Index')
    plt.ylabel('Predicted Probability')
    plt.tight_layout()
    plt.savefig('decision_tree_predicted_probabilities.png')
    plt.close()

    # Export line plot to Excel
    wb_dt = Workbook()
    ws_dt = wb_dt.active
    ws_dt.title = "Decision Tree Prediction"
    ws_dt['A1'] = 'Predicted Probability of Cancer Cell Presence'
    img_dt = XLImage('decision_tree_predicted_probabilities.png')
    ws_dt.add_image(img_dt, 'B3')
    wb_dt.save('decision_tree_prediction_plot.xlsx')

    # Save the model
    with open('decision_tree_model.pkl', 'wb') as f:
        pickle.dump(dt_model, f)

    print("✅ Model saved as decision_tree_model.pkl")


    #testing of the model 
    from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report

    # Assuming you have:
    # logistic_model - your trained model
    # X_test, y_test - your test dataset and labels
    logistic_model = LogisticRegression()
    logistic_model.fit(X_train, y_train)
    with open('logistic_model.pkl', 'wb') as f:
        pickle.dump(logistic_model, f)
    with open('logistic_model.pkl', 'rb') as f:
        logistic_model = pickle.load(f)
    
    # Make predictions
    y_pred = logistic_model.predict(X_test)

    # Calculate metrics
    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred)    # for binary classification, default pos_label=1
    recall = recall_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)

    print(f"Accuracy:  {accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall:    {recall:.4f}")
    print(f"F1 Score:  {f1:.4f}")

    # Optional: detailed classification report
    print("\nClassification Report:\n", classification_report(y_test, y_pred))


